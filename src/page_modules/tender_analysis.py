"""
Use Case 1 - Analyzing tender documents & fill out internal tender list.

This page lets users:
- Upload tender documents
- AI automatically extracts key information (German translation + fields)
- Fills out internal tender list template (first sheet)
- Export results as Excel
"""

from io import BytesIO
import hashlib

import pandas as pd
import streamlit as st
from openpyxl.utils import get_column_letter

from config.settings import azure_config
from utils.ai_analyzer import analyze_tender_document, analyze_tender_with_fields
from utils.pdf_processor import extract_text_from_pdf, get_text_length_info


def render_tender_analysis_page():
    st.title("Tender Document Analysis")
    azure_config.show_credentials_warning()

    st.markdown(
        """
        **Workflow:**
        1) Upload tender PDFs
        2) Translate and extract key fields in German
        3) Auto-fill your internal tender list (1st sheet of an XLSX template)
        """
    )

    st.markdown("---")

    col1, col2 = st.columns(2)
    
    with col1:
        tender_files = st.file_uploader(
            "Upload tender documents (PDF)",
            type=["pdf"],
            accept_multiple_files=True,
            help="You can upload multiple PDF tender documents."
        )

    with col2:
        tender_template = st.file_uploader(
            "Upload internal tender list template (XLSX)",
            type=["xlsx"],
            key="tender_template_upload",
            help="The first sheet will be filled with extracted data."
        )

    # Initialize from session to keep context across reruns (e.g., after downloads)
    template_structure = st.session_state.get("tender_template_structure")
    desired_fields = st.session_state.get("tender_desired_fields")

    run_analysis = st.button("Analyze tender documents", use_container_width=True)

    if run_analysis:
        if not tender_files:
            st.warning("Please upload at least one tender PDF.")
            return

        if not tender_template:
            st.warning("Please upload an internal tender list template (XLSX).")
            return

        results = []
        progress = st.progress(0.0)
        status_placeholder = st.empty()

        # If a template is provided, extract field labels from column E (form-style)
        try:
            template_bytes = tender_template.getvalue()
            template_hash = hashlib.md5(template_bytes).hexdigest()
            
            # Cache template parsing to avoid re-parsing same file
            if st.session_state.get("tender_template_hash") != template_hash:
                template_structure = _parse_form_template(tender_template)
                st.session_state["tender_template_structure"] = template_structure
                st.session_state["tender_template_hash"] = template_hash
            else:
                template_structure = st.session_state.get("tender_template_structure")
            
            desired_fields = template_structure['field_names']
            st.info(f"Extracted {len(desired_fields)} fields from template: {', '.join(desired_fields[:5])}...")
            st.session_state["tender_desired_fields"] = desired_fields
        except Exception as e:
            st.warning(f"Could not read template: {e}.")
            return

        for idx, tender_file in enumerate(tender_files):
            status_placeholder.info(f"Processing {tender_file.name} ...")
            
            # Cache PDF extraction to avoid re-processing same file
            pdf_cache_key = f"pdf_cache_{tender_file.name}_{tender_file.size}"
            if pdf_cache_key not in st.session_state:
                extracted_text = extract_text_from_pdf(tender_file)
                st.session_state[pdf_cache_key] = extracted_text
            else:
                extracted_text = st.session_state[pdf_cache_key]
            
            length_info = get_text_length_info(extracted_text)
            if desired_fields:
                analysis = analyze_tender_with_fields(extracted_text, desired_fields)
            else:
                analysis = analyze_tender_document(extracted_text)
            success = "error" not in analysis
            results.append({
                "file_name": tender_file.name,
                "analysis": analysis,
                "success": success,
                "text_length": length_info.get("length", len(extracted_text))
            })
            progress.progress((idx + 1) / len(tender_files))

        status_placeholder.success("Analysis completed.")
        st.session_state["tender_results"] = results

    results = st.session_state.get("tender_results", [])

    if results:
        success_count = sum(1 for item in results if item.get("success"))
        st.markdown(f"**Processed:** {success_count}/{len(results)} documents")

       
        template_structure = st.session_state.get("tender_template_structure")
        if template_structure and results:
            successful_results = [r for r in results if r.get("success")]
            if successful_results:
                merged_analysis = _merge_tender_analyses(successful_results)
                merged_result = {
                    "file_name": f"Tender Package ({len(successful_results)} file(s))",
                    "analysis": merged_analysis,
                    "success": True
                }

                # Display merged fields in table format (Feld | Wert)
                extracted = merged_analysis.get("extracted", {})
                field_items = list(extracted.items())
                df = pd.DataFrame(field_items)
                df.columns = ["Feld", "Wert"]

                def highlight_row(row):
                    Wert = str(row.get("Wert", ""))
                    if "Nicht angegeben" in Wert:
                        return ["background-color: rgba(220, 53, 69, 0.3); border-left: 3px solid #dc3545"] * len(row)
                    return [""] * len(row)
                
                styled = df.style.apply(highlight_row, axis=1)
                st.dataframe(styled, use_container_width=True, hide_index=True)
                
                workbook = _fill_form_template(merged_result, tender_template, template_structure)
                st.download_button(
                    f"📥 Download: Tender Package (Filled)",
                    data=workbook.getvalue(),
                    file_name="tender_package_filled.xlsx",
                    mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                    key="download_merged",
                    help=f"Filled tender form from {len(successful_results)} file(s).",
                    use_container_width=True
                )
        else:
            # Fallback: generic table-style export
            prepared_rows = [_build_tender_row(r) for r in results if r.get("success")]
            if prepared_rows:
                workbook = _build_tender_workbook(prepared_rows, None)
                st.download_button(
                    "Download filled tender list (Excel)",
                    data=workbook.getvalue(),
                    file_name="tender_list_filled.xlsx",
                    mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                    use_container_width=True,
                    help="First sheet contains the auto-filled entries."
                )



def _merge_tender_analyses(results: list) -> dict:
    """Merge analyses from multiple documents (same tender set).
    
    Priority order for field values:
    1. First non-empty extracted value wins
    2. Concatenate German summaries from all documents
    """
    if not results:
        return {"extracted": {}, "german_summary": "", "notes": ""}
    
    # Start with first result's extracted fields
    merged_extracted = dict(results[0].get("analysis", {}).get("extracted", {}))
    
    # Fill in missing values from subsequent documents
    for result in results[1:]:
        extracted = result.get("analysis", {}).get("extracted", {})
        for field, value in extracted.items():
            if not merged_extracted.get(field) or merged_extracted[field] == "Nicht angegeben":
                merged_extracted[field] = value
    
    # Concatenate German summaries
    summaries = []
    for result in results:
        summary = result.get("analysis", {}).get("german_summary", "").strip()
        if summary:
            summaries.append(f"[{result.get('file_name', 'Document')}] {summary}")
    
    merged_summary = "\n\n".join(summaries) if summaries else ""
    
    # Concatenate notes
    notes_list = []
    for result in results:
        note = result.get("analysis", {}).get("notes", "").strip()
        if note:
            notes_list.append(f"[{result.get('file_name', 'Document')}] {note}")
    
    merged_notes = "\n\n".join(notes_list) if notes_list else ""
    
    return {
        "extracted": merged_extracted,
        "german_summary": merged_summary,
        "notes": merged_notes
    }



def _parse_form_template(template_file) -> dict:
    """Parse form-style template: extract field names from column E and their row positions."""
    base_bytes = template_file.getvalue()
    excel_file = pd.ExcelFile(BytesIO(base_bytes))
    first_sheet = excel_file.sheet_names[0]
    df = pd.read_excel(excel_file, sheet_name=first_sheet, header=None)

    # Extract field names from column E (index 4) where values end with ':'
    field_map = {}
    field_names = []

    for idx, row in df.iterrows():
        if len(row) > 4 and pd.notna(row[4]):
            field_label = str(row[4]).strip()
            if field_label.endswith(':'):
                field_label_clean = field_label.rstrip(':')
                field_map[field_label_clean] = {
                    'row': idx,
                    'label_col': 4,  # Column E
                    'value_col': 5    # Column F for values
                }
                field_names.append(field_label_clean)

    return {
        'field_names': field_names,
        'field_map': field_map,
        'sheet_name': first_sheet,
        'template_bytes': base_bytes
    }


def _fill_form_template(result: dict, template_file, template_structure: dict) -> BytesIO:
    """Fill a form-style template with extracted values for a single tender document."""
    from openpyxl import load_workbook
    
    analysis = result.get("analysis", {})
    extracted = analysis.get("extracted", {})
    
    # Load template workbook
    workbook = load_workbook(BytesIO(template_structure['template_bytes']))
    sheet = workbook[template_structure['sheet_name']]
    
    # Fill each field
    field_map = template_structure['field_map']
    for field_name, position in field_map.items():
        value = extracted.get(field_name, "Nicht angegeben")
        # Excel uses 1-based indexing, pandas uses 0-based
        row_num = position['row'] + 1
        col_letter = get_column_letter(position['value_col'] + 1)  # Convert to Excel column letter
        cell = sheet[f"{col_letter}{row_num}"]
        cell.value = value
    
    # Optionally fill project description if row 4 exists
    summary = analysis.get("german_summary", "")
    if summary:
        try:
            sheet['C4'] = summary  # Adjust as needed
        except:
            pass
    
    # Save to BytesIO
    output = BytesIO()
    workbook.save(output)
    output.seek(0)
    return output


def _build_tender_row(result: dict) -> dict:
    analysis = result.get("analysis", {})
    fields = analysis.get("tender_fields", {})
    return {
        "source_file": result.get("file_name", ""),
        "customer": fields.get("customer", ""),
        "project_title": fields.get("project_title", ""),
        "procedure": fields.get("procedure", ""),
        "reference_number": fields.get("reference_number", ""),
        "submission_deadline": fields.get("submission_deadline", ""),
        "questions_deadline": fields.get("questions_deadline", ""),
        "contract_start": fields.get("contract_start", ""),
        "contract_end": fields.get("contract_end", ""),
        "estimated_value": fields.get("estimated_value", ""),
        "country": fields.get("country", ""),
        "language": fields.get("language", ""),
        "cpv_codes": ", ".join(fields.get("cpv_codes", [])),
        "notes": fields.get("notes", ""),
        "summary": analysis.get("german_summary", ""),
        "risks": "; ".join(analysis.get("risks", [])),
        "scope": "; ".join(analysis.get("key_requirements", [])),
    }


def _build_tender_workbook(rows: list, template_upload) -> BytesIO:
    if not rows:
        return BytesIO()

    if template_upload is not None:
        base_bytes = template_upload.getvalue()
        excel_file = pd.ExcelFile(BytesIO(base_bytes))
        sheets = {name: pd.read_excel(excel_file, sheet_name=name) for name in excel_file.sheet_names}
        first_sheet = excel_file.sheet_names[0]
        sheets[first_sheet] = _append_rows_to_df(sheets[first_sheet], rows)
    else:
        default_cols = [
            "Source File", "Customer", "Project Title", "Procedure", "Reference Number",
            "Submission Deadline", "Questions Deadline", "Contract Start", "Contract End",
            "Estimated Value", "Country", "Language", "CPV Codes", "Scope/Requirements",
            "Risks", "Summary", "Notes"
        ]
        sheets = {"Tenders": pd.DataFrame([_map_row_to_default(r, default_cols) for r in rows])}

    output = BytesIO()
    with pd.ExcelWriter(output, engine="openpyxl") as writer:
        for name, df in sheets.items():
            df.to_excel(writer, sheet_name=name, index=False)
    output.seek(0)
    return output


def _append_rows_to_df(df: pd.DataFrame, rows: list) -> pd.DataFrame:
    mapped = [_map_row_to_columns(row, list(df.columns)) for row in rows]
    return pd.concat([df, pd.DataFrame(mapped)], ignore_index=True)


def _map_row_to_columns(row: dict, columns: list) -> dict:
    column_values = {col: "" for col in columns}
    mapping = {
        "customer": ["auftraggeber", "kunde", "customer", "client", "buyer"],
        "project_title": ["projekt", "project", "title", "leistung"],
        "procedure": ["verfahren", "procedure"],
        "reference_number": ["referenz", "aktenzeichen", "reference"],
        "submission_deadline": ["frist", "submission", "angebot", "abgabe", "deadline"],
        "questions_deadline": ["fragen", "question", "clarification"],
        "contract_start": ["start", "beginn"],
        "contract_end": ["ende", "end", "laufzeit"],
        "estimated_value": ["wert", "value", "budget"],
        "country": ["land", "country"],
        "language": ["sprache", "language"],
        "cpv_codes": ["cpv"],
        "scope": ["scope", "anforder", "requirement", "leistung"],
        "risks": ["risiko", "risk"],
        "summary": ["summary", "zusammenfassung"],
        "notes": ["note", "hinweis", "bemerk"],
        "source_file": ["file", "datei", "quelle"]
    }

    for col in columns:
        col_lower = col.lower()
        value = ""
        for key, hints in mapping.items():
            if any(hint in col_lower for hint in hints):
                value = row.get(key, "")
                break
        column_values[col] = value

    return column_values


def _map_row_to_default(row: dict, columns: list) -> dict:
    ordered = []
    for col in columns:
        col_lower = col.lower()
        if "source" in col_lower:
            ordered.append(row.get("source_file", ""))
        elif "customer" in col_lower:
            ordered.append(row.get("customer", ""))
        elif "project" in col_lower:
            ordered.append(row.get("project_title", ""))
        elif "procedure" in col_lower:
            ordered.append(row.get("procedure", ""))
        elif "reference" in col_lower:
            ordered.append(row.get("reference_number", ""))
        elif "submission" in col_lower:
            ordered.append(row.get("submission_deadline", ""))
        elif "question" in col_lower:
            ordered.append(row.get("questions_deadline", ""))
        elif "start" in col_lower:
            ordered.append(row.get("contract_start", ""))
        elif "end" in col_lower:
            ordered.append(row.get("contract_end", ""))
        elif "value" in col_lower:
            ordered.append(row.get("estimated_value", ""))
        elif "country" in col_lower:
            ordered.append(row.get("country", ""))
        elif "language" in col_lower:
            ordered.append(row.get("language", ""))
        elif "cpv" in col_lower:
            ordered.append(row.get("cpv_codes", ""))
        elif "scope" in col_lower or "requirement" in col_lower:
            ordered.append(row.get("scope", ""))
        elif "risk" in col_lower:
            ordered.append(row.get("risks", ""))
        elif "summary" in col_lower:
            ordered.append(row.get("summary", ""))
        elif "note" in col_lower:
            ordered.append(row.get("notes", ""))
        else:
            ordered.append("")
    return dict(zip(columns, ordered))
