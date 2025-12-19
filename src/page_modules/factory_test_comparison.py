"""
Use Case 3 - Comparison of factory test certificates against specifications.

This page lets users:
- Upload all PDFs (specifications and certificates) together
- AI automatically identifies which are specs vs certificates
- AI extracts parameters and performs intelligent comparison
- Displays comparison results with tolerance checks
- Export results to CSV
"""

from __future__ import annotations

import json
from typing import Dict, Any
from io import BytesIO
import os
import re
from pathlib import Path
from datetime import datetime
from openpyxl import load_workbook
from openpyxl.styles import Font


import pandas as pd
import streamlit as st

from utils.pdf_processor import extract_text_from_pdf


# --------------
# Page rendering
# --------------


def render_factory_test_comparison_page():
    
    st.title("Factory Certificate Comparison")
    st.markdown(
        """
        **Simple AI-powered certificate comparison:**
        - Drop all your PDFs (specifications and certificates) in one upload
        - AI will automatically identify specs vs certificates
        - Get instant comparison table with tolerance analysis
        """
    )

    st.markdown("---")

    # Single file uploader for all PDFs
    uploaded_files = st.file_uploader(
        "Upload all PDFs (specifications and certificates)",
        type=["pdf"],
        accept_multiple_files=True,
        key="all_pdfs_upload",
        help="Drop both specification PDFs and certificate PDFs here - AI will figure out which is which",
    )

    if st.button("Analyze & Compare", width="stretch"):
        _process_ai_comparison(uploaded_files )
    
    # Display results from session state if available
    if "comparison_result" in st.session_state and st.session_state.comparison_result:
        _display_smart_comparison_results(st.session_state.comparison_result)

def _process_ai_comparison(uploaded_files):
    """Process all PDFs and let AI do the identification and comparison."""
    progress = st.progress(0)
    status = st.empty()
    
    # Step 1: Extract text from all PDFs
    status.text("Extracting text from all PDFs...")
    file_texts = {}
    
    for idx, file in enumerate(uploaded_files):
        try:
            text = extract_text_from_pdf(file)
            file_texts[file.name] = text
            progress.progress((idx + 1) / (len(uploaded_files) + 1))
        except Exception as e:
            st.error(f"Failed to extract text from {file.name}: {e}")
            continue
    
    if len(file_texts) < 2:
        st.error("Could not extract text from enough files. Need at least 2 readable PDFs.")
        return
    
    # Step 2: Send all to AI for intelligent comparison
    status.text("AI analyzing and comparing all documents...")
    comparison_result = _ai_smart_compare(file_texts)
    progress.progress(1.0)
    
    if comparison_result.get("error"):
        st.error(f"AI comparison failed: {comparison_result['error']}")
        return
    
    status.text("Analysis complete! ✅")
    
    # Store results in session state for persistence across reruns
    st.session_state.comparison_result = comparison_result
    
    # Rerun to display results from session state
    st.rerun()



def _ai_smart_compare(file_texts: Dict[str, str]) -> Dict[str, Any]:
    """Let AI identify specs vs certificates and perform comparison."""
    from config.settings import azure_config
    
    if not azure_config.client:
        return {"error": "Azure OpenAI not configured"}
    
    # Build context with all files
    files_context = ""
    for idx, (filename, text) in enumerate(file_texts.items(), 1):
        files_context += f"\n\n=== DOCUMENT {idx}: {filename} ===\n{text[:8000]}\n"  # Limit each doc
    
    prompt = f"""
You are analyzing multiple technical documents. Your task:

1. **IDENTIFY** which documents are specifications vs certificates/test reports
2. **EXTRACT** all parameters from specifications (with tolerances, min/max, nominal values)
3. **EXTRACT** all measured values from certificates
4. **COMPARE** measured values against specification tolerances
5. **RETURN** a complete comparison table

Documents to analyze:
{files_context}

Return ONLY valid JSON with this structure:
{{
  "identified_specs": ["list of spec filenames"],
  "identified_certificates": ["list of certificate filenames"],
  "comparisons": [
    {{
      "parameter": "parameter name",
      "unit": "unit of measurement",
      "spec_min": number or null,
      "spec_max": number or null,
      "spec_nominal": number or null,
      "measured_value": number or null,
      "measured_from": "certificate filename",
      "status": "OK" | "OUT" | "MISSING" | "NO_SPEC",
      "deviation": "description of deviation if OUT"
    }}
  ],
  "summary": "Brief summary of comparison results"
}}

Status definitions:
- OK: Measured value within specification tolerances
- OUT: Measured value outside tolerances
- MISSING: Parameter in spec but no measurement found
- NO_SPEC: Measurement found but no specification for it

DO NOT use markdown. Return raw JSON only.
"""
    
    try:
        response = azure_config.client.chat.completions.create(
            model=azure_config.deployment_name,
            messages=[{"role": "user", "content": prompt}],
        )
        response_text = response.choices[0].message.content.strip()
        
        # Clean markdown if present
        if "```json" in response_text:
            response_text = response_text.split("```json")[1].split("```")[0].strip()
        elif "```" in response_text:
            response_text = response_text.split("```")[1].split("```")[0].strip()
        
        data = json.loads(response_text)
        return data
    except Exception as e:
        return {"error": str(e)}


def _display_smart_comparison_results(result: Dict[str, Any]):
    """Display the AI comparison results."""
    st.markdown("## 📊 Analysis Results")
    
    def _make_export_basename(res: Dict[str, Any]) -> str:
        specs = res.get("identified_specs", []) or []
        certs = res.get("identified_certificates", []) or []
        names = [Path(n).stem for n in (specs + certs) if isinstance(n, str) and n]
        if not names:
            base = "analysis"
        elif len(specs) == 1:
            base = Path(specs[0]).stem
        elif len(certs) == 1:
            base = Path(certs[0]).stem
        else:
            cp = os.path.commonprefix(names)
            cp = re.sub(r"[-_\s.]+$", "", cp)
            if len(cp) >= 3:
                base = cp
            else:
                token_lists = [re.findall(r"[A-Za-z0-9]+", n) for n in names]
                counts = {}
                for tokens in token_lists:
                    seen = set()
                    for t in tokens:
                        if t in seen:
                            continue
                        seen.add(t)
                        counts[t] = counts.get(t, 0) + 1
                threshold = max(1, int(0.6 * len(names)))
                first_tokens = token_lists[0] if token_lists else []
                chosen = [t for t in first_tokens if counts.get(t, 0) >= threshold]
                base = "-".join(chosen) if chosen else names[0]
        base = re.sub(r"[^A-Za-z0-9._-]", "-", base)
        base = re.sub(r"-+", "-", base).strip("-._")
        date_tag = datetime.now().strftime("%Y%m%d")
        return f"{base}-{date_tag}" if base else f"analysis-{date_tag}"
    
    # Show document identification
    with st.expander("🔍 Document Identification", expanded=False):
        col1, col2 = st.columns(2)
        with col1:
            st.markdown("**Specifications:**")
            for spec in result.get("identified_specs", []):
                st.write(f"• {spec}")
        with col2:
            st.markdown("**Certificates:**")
            for cert in result.get("identified_certificates", []):
                st.write(f"• {cert}")
    
    # Summary
    if result.get("summary"):
        st.info(f"**Summary:** {result['summary']}")
    
    # Comparison table
    comparisons = result.get("comparisons", [])
    if not comparisons:
        st.warning("No comparisons found.")
        return
    
    df = pd.DataFrame(comparisons)
    
    # Calculate metrics
    ok_count = (df["status"].str.upper() == "OK").sum()
    out_count = (df["status"].str.upper() == "OUT").sum()
    missing_count = (df["status"].str.upper() == "MISSING").sum()
    
    c1, c2, c3 = st.columns(3)
    with c1:
        st.metric("Within Tolerance", int(ok_count))
    with c2:
        st.metric("Out of Tolerance", int(out_count))
    with c3:
        st.metric("Missing Measurements", int(missing_count))
    
    st.markdown("---")
    st.markdown("### 📋 Comparison Table")
    
    # Color coding - using colors that work in both light and dark mode
    def highlight_row(row):
        status = str(row.get("status", "")).upper()
        if status == "OUT":
            return ["background-color: rgba(220, 53, 69, 0.3); border-left: 3px solid #dc3545"] * len(row)
        if status == "MISSING":
            return ["background-color: rgba(255, 193, 7, 0.3); border-left: 3px solid #ffc107"] * len(row)
        if status == "OK":
            return ["background-color: rgba(40, 167, 69, 0.3); border-left: 3px solid #28a745"] * len(row)
        return [""] * len(row)
    
    styled = df.style.apply(highlight_row, axis=1)
    st.dataframe(styled, width="stretch", hide_index=True)
    
    # Export
    st.markdown("### 📥 Export Results")
    export_base = _make_export_basename(result)
    
    def excel_highlight_row(row):
        status = str(row.get("status", "")).upper()
        if status == "OUT":
            return ["background-color: #dc3545; border-left: 3px solid #dc3545"] * len(row)
        if status == "MISSING":
            return ["background-color: #ffc107; border-left: 3px solid #ffc107"] * len(row)
        if status == "OK":
            return ["background-color: #28a745; border-left: 3px solid #28a745"] * len(row)
        return [""] * len(row)

    excel_buffer = BytesIO()
    try:
        excel_styled = df.style.apply(excel_highlight_row, axis=1)
        excel_styled.to_excel(excel_buffer, engine="openpyxl", index=False)

        # Post-process with openpyxl to enforce a font that renders special characters well
        excel_buffer.seek(0)
        wb = load_workbook(excel_buffer)
        ws = wb.active
        default_font = Font(name="Calibri")
        for row in ws.iter_rows(min_row=1, max_row=ws.max_row, min_col=1, max_col=ws.max_column):
            for cell in row:
                cell.font = default_font

        final_buffer = BytesIO()
        wb.save(final_buffer)
        st.download_button(
            label="Download Comparison Excel",
            data=final_buffer.getvalue(),
            file_name=f"{export_base}_comparison.xlsx",
            mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        )
    except Exception as e:
        st.warning(f"Excel export unavailable: {e}")
